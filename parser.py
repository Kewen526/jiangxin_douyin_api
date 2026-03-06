"""
Excel 解析模块 — 将下载的 xlsx 文件解析为上传 API 所需的字典列表
"""

from openpyxl import load_workbook

import config


def parse_xlsx(file_path):
    """
    解析 xlsx 文件，返回 (rows, raw_row_count)。

    - rows: list[dict]，每条数据已映射为 API 字段名，可直接用于上传
    - raw_row_count: 数据行数（不含标题行），用于判断是否为空数据

    如果只有标题行（raw_row_count == 0），rows 为空列表。
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return [], 0

    # 第一行为标题
    header_row = all_rows[0]
    data_row_list = all_rows[1:]

    print(f"  表头：{list(header_row)}")
    print(f"  总行数：{len(all_rows)}（标题 1 行 + 数据 {len(data_row_list)} 行）")

    # 建立列索引：Excel 列名 → (位置索引, API 字段名)
    col_map = {}  # {col_index: api_field_name}
    for idx, cell_value in enumerate(header_row):
        if cell_value is None:
            continue
        col_name = str(cell_value).strip()
        api_field = config.EXCEL_COLUMN_MAPPING.get(col_name)
        if api_field:
            col_map[idx] = api_field

    if not col_map:
        print(f"  警告：未匹配到任何已知列名")
        return [], 0

    print(f"  匹配到 {len(col_map)} 个字段列：{list(col_map.values())}")

    # 解析数据行
    data_rows = []
    for row in data_row_list:
        record = {}
        for col_idx, api_field in col_map.items():
            value = row[col_idx] if col_idx < len(row) else None
            # 金额字段转为数值
            if api_field in config.NUMERIC_FIELDS:
                value = _to_number(value)
            else:
                value = str(value).strip() if value is not None else ""
            record[api_field] = value
        data_rows.append(record)

    raw_row_count = len(data_row_list)
    print(f"  解析完成：{raw_row_count} 行数据")
    return data_rows, raw_row_count


def _to_number(value):
    """将单元格值安全转换为数值"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return round(value, 2)
    s = str(value).strip().replace(",", "")
    if not s or s == "-":
        return 0
    try:
        return round(float(s), 2)
    except ValueError:
        return 0
